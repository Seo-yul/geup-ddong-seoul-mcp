"""서울시 공중화장실 데이터 계층.

스마트서울맵(theme_id=100106) 테마 갤러리에서 제공하는 공중화장실 데이터를
다운로드 -> 압축 해제 -> 파싱 -> 메모리 적재하고, 조회용 헬퍼를 제공한다.

다운로드 엔드포인트(로그인 불필요, POST):
  https://map.seoul.go.kr/smgis/webs/contents/contents.do
  ?mode=downloadGeomFormatXLSX&down_type=xlsx&lan_type=KOR&conts_type=A&theme_id=100106
응답: contents.zip  ->  내부에 contents.xlsx(데이터), coord.xlsx(점 데이터라 미사용)
"""

from __future__ import annotations

import io
import json
import math
import re
import threading
import zipfile
from datetime import datetime, time
from pathlib import Path
from typing import Any, Iterable, Optional

import requests
from openpyxl import load_workbook

try:
    from zoneinfo import ZoneInfo

    KST = ZoneInfo("Asia/Seoul")
except Exception:  # pragma: no cover - fallback for환경 차이
    from datetime import timedelta, timezone

    KST = timezone(timedelta(hours=9))

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------

THEME_ID = "100106"
DOWNLOAD_URL = "https://map.seoul.go.kr/smgis/webs/contents/contents.do"
DOWNLOAD_PARAMS = {
    "mode": "downloadGeomFormatXLSX",
    "down_type": "xlsx",
    "lan_type": "KOR",
    "conts_type": "A",
    "theme_id": THEME_ID,
}
# 일부 공공 서버는 기본 UA를 차단하므로 브라우저 UA + Referer를 함께 보낸다.
_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Referer": f"https://map.seoul.go.kr/smgis2/themeGallery/detail?theme_id={THEME_ID}",
}

DEFAULT_CACHE_DIR = Path(__file__).resolve().parent.parent / "data_cache"
INNER_XLSX_NAME = "contents.xlsx"

# 데이터셋 라벨(설명용)
THEME_NAME = "서울시 공중화장실"

# ---------------------------------------------------------------------------
# 컬럼 매핑 (헤더 1행의 첫 줄 텍스트 기준)
# ---------------------------------------------------------------------------

# 고정 컬럼: 헤더 첫 줄 텍스트 -> 내부 필드명
_FIXED_COLUMNS = {
    "콘텐츠 ID": "content_id",
    "사용유무": "in_use_raw",
    "콘텐츠명": "name",
    "서브카테고리 명": "subcategory",
    "구명": "district",
    "새주소[도로명 주소]": "road_address",
    "지번주소": "jibun_address",
    "키워드": "keyword",
    "좌표[X]": "lon",
    "좌표[Y]": "lat",
    "전화번호": "phone",
    "웹URL": "web_url",
}

# "상세 제목N" / "상세 내용N" 컬럼 패턴
_DETAIL_TITLE_RE = re.compile(r"상세\s*제목\s*(\d+)")
_DETAIL_CONTENT_RE = re.compile(r"상세\s*내용\s*(\d+)")

# 시간 범위 추출: 08:00~21:00 / 8:00-21:00 등
_TIME_RANGE_RE = re.compile(r"(\d{1,2})\s*:\s*(\d{2})\s*[~\-–]\s*(\d{1,2})\s*:\s*(\d{2})")


def _first_line(text: Any) -> str:
    if text is None:
        return ""
    return str(text).split("\n")[0].strip()


def _split_multi(value: Any) -> list[str]:
    """'남자|여자|' 같은 파이프 구분 다중값을 리스트로."""
    if value is None:
        return []
    parts = [p.strip() for p in str(value).split("|")]
    return [p for p in parts if p]


def _to_float(value: Any) -> Optional[float]:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(str(value).strip())
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# 개방시간 파싱
# ---------------------------------------------------------------------------


def parse_open_hours(raw: Any, subcategory: Any = None) -> dict:
    """개방시간 문자열을 구조화한다.

    반환: {"raw", "is_24h", "ranges": [[start_min, end_min], ...], "note"}
    - 상시 / 24시간 / 00:00~24:00 => is_24h True
    - "정시(08:00~21:00)" => ranges=[[480, 1260]]
    - 파싱 불가 => ranges=[] (is_open_now 시 None 반환)
    """
    text = "" if raw is None else str(raw).strip().rstrip("|").strip()
    sub = "" if subcategory is None else str(subcategory)
    note_bits: list[str] = []

    is_24h = False
    if any(k in (text + sub) for k in ("상시", "24시간", "24時", "00:00~24:00", "0:00~24:00")):
        is_24h = True

    ranges: list[list[int]] = []
    for m in _TIME_RANGE_RE.finditer(text):
        sh, sm, eh, em = (int(m.group(i)) for i in range(1, 5))
        start = sh * 60 + sm
        end = eh * 60 + em
        # 24:00 표기는 1440분으로 처리
        if end == 0 and start != 0:
            end = 24 * 60
        ranges.append([start, end])

    # 00:00~24:00 전일 개방도 24h로 간주
    for s, e in ranges:
        if s <= 0 and e >= 24 * 60:
            is_24h = True

    if any(k in text for k in ("평일", "주말", "공휴일", "토요일", "일요일")):
        note_bits.append("요일/공휴일별 차이가 있을 수 있음(원문 확인 권장)")
    if not ranges and not is_24h and text:
        note_bits.append("개방시간 형식을 자동 해석하지 못함(원문 확인 권장)")

    return {
        "raw": text,
        "is_24h": is_24h,
        "ranges": ranges,
        "note": "; ".join(note_bits) or None,
    }


def is_open_now(parsed: dict, now: Optional[datetime] = None) -> Optional[bool]:
    """현재(KST) 개방 여부. 판단 불가 시 None."""
    if parsed.get("is_24h"):
        return True
    ranges = parsed.get("ranges") or []
    if not ranges:
        return None
    if now is None:
        now = datetime.now(KST)
    cur = now.hour * 60 + now.minute
    for start, end in ranges:
        if start == end:
            return True  # 동일 표기는 종일로 간주
        if start < end:
            if start <= cur < end:
                return True
        else:  # 자정을 넘기는 구간 (예: 22:00~06:00)
            if cur >= start or cur < end:
                return True
    return False


# ---------------------------------------------------------------------------
# 거리 계산
# ---------------------------------------------------------------------------


def haversine_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """두 좌표 사이의 대권 거리(미터)."""
    r = 6371000.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlmb / 2) ** 2
    return 2 * r * math.asin(min(1.0, math.sqrt(a)))


# ---------------------------------------------------------------------------
# 파싱
# ---------------------------------------------------------------------------


def _detail_key(title: str) -> Optional[str]:
    """상세 제목 텍스트를 구조화 필드명으로 매핑."""
    t = title.replace(" ", "")
    if "개방주체" in t:
        return "open_subject"
    if "개방시간" in t:
        return "open_hours"
    if "휴관" in t:
        return "closed_days"
    if "장애인" in t:  # 장애인화장실구분 / 장애인화장구분
        return "disabled_toilet"
    if "화장실구분" in t:
        return "toilet_types"
    if "편의시설" in t:
        return "amenities"
    if "안전시설" in t:
        return "safety"
    if "건물용도" in t:
        return "building_use"
    if "관리기관" in t:
        return "manager"
    return None


# 다중값(파이프) 으로 다루는 구조화 필드
_MULTI_FIELDS = {"closed_days", "toilet_types", "disabled_toilet", "amenities", "safety"}


def parse_xlsx(xlsx_bytes: bytes) -> list[dict]:
    """contents.xlsx 바이트를 화장실 레코드 리스트로 파싱한다."""
    # 스마트서울맵이 내려주는 xlsx는 시트의 <dimension>이 'A1'로 잘못 박혀 있어
    # read_only 스트리밍 모드로 열면 1행 1열만 읽혀 0건이 된다(reset_dimensions로도
    # 복구 안 됨). 전체 로드 모드로 열어 4,500여 행을 모두 읽는다(파일 약 1.8MB, 피크 ~170MB).
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    header = next(rows)
    headers = [_first_line(h) for h in header]

    fixed_idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h in _FIXED_COLUMNS and _FIXED_COLUMNS[h] not in fixed_idx:
            fixed_idx[_FIXED_COLUMNS[h]] = i

    # 상세 제목/내용 컬럼 짝 찾기
    title_cols: dict[str, int] = {}
    content_cols: dict[str, int] = {}
    for i, h in enumerate(headers):
        mt = _DETAIL_TITLE_RE.search(h)
        mc = _DETAIL_CONTENT_RE.search(h)
        if mt:
            title_cols[mt.group(1)] = i
        elif mc:
            content_cols[mc.group(1)] = i
    detail_pairs = [
        (title_cols[k], content_cols[k]) for k in title_cols if k in content_cols
    ]

    def cell(row, field):
        idx = fixed_idx.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    records: list[dict] = []
    for row in rows:
        if row is None:
            continue
        content_id = cell(row, "content_id")
        name = cell(row, "name")
        if content_id is None and name is None:
            continue

        # 상세 제목->내용 딕셔너리
        details_raw: dict[str, str] = {}
        for ti, ci in detail_pairs:
            title = _first_line(row[ti]) if ti < len(row) else ""
            value = row[ci] if ci < len(row) else None
            if title and value not in (None, ""):
                # 원본은 값 끝에 구분자 '|'가 붙는다(예: "공공시설|"). 끝 파이프만 정리.
                cleaned = str(value).strip().rstrip("|").strip()
                if cleaned:
                    details_raw[title] = cleaned

        rec: dict[str, Any] = {
            "content_id": (str(content_id).strip() if content_id is not None else None),
            "name": (str(name).strip() if name is not None else None),
            "subcategory": _opt_str(cell(row, "subcategory")),
            "in_use": str(cell(row, "in_use_raw")).strip().upper() == "Y",
            "district": _opt_str(cell(row, "district")),
            "road_address": _opt_str(cell(row, "road_address")),
            "jibun_address": _opt_str(cell(row, "jibun_address")),
            "keyword": _opt_str(cell(row, "keyword")),
            "phone": _opt_str(cell(row, "phone")),
            "web_url": _opt_str(cell(row, "web_url")),
            "lat": _to_float(cell(row, "lat")),
            "lon": _to_float(cell(row, "lon")),
            "details": details_raw,
        }

        # 구조화 상세 필드
        structured: dict[str, Any] = {}
        for title, value in details_raw.items():
            key = _detail_key(title)
            if key is None:
                continue
            if key in _MULTI_FIELDS:
                structured[key] = _split_multi(value)
            else:
                structured[key] = value
        rec.update(structured)

        # 개방시간 파생
        oh = parse_open_hours(rec.get("open_hours"), rec.get("subcategory"))
        rec["open_hours_parsed"] = oh

        records.append(rec)

    wb.close()
    return records


def _opt_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


# ---------------------------------------------------------------------------
# 다운로드 / 압축 해제
# ---------------------------------------------------------------------------


def download_zip(timeout: int = 60) -> bytes:
    """contents.zip 바이트를 다운로드(POST, 인증 불필요)."""
    resp = requests.post(
        DOWNLOAD_URL, params=DOWNLOAD_PARAMS, headers=_HEADERS, timeout=timeout
    )
    resp.raise_for_status()
    content = resp.content
    if content[:2] != b"PK":
        raise ValueError(
            "다운로드 응답이 ZIP이 아닙니다. 엔드포인트/파라미터를 확인하세요."
        )
    return content


def extract_inner_xlsx(zip_bytes: bytes) -> bytes:
    """contents.zip 에서 contents.xlsx 바이트를 추출."""
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()
        target = INNER_XLSX_NAME
        if target not in names:
            target = next((n for n in names if n.lower().endswith(".xlsx")), None)
            if target is None:
                raise ValueError(f"zip 안에 xlsx가 없습니다: {names}")
        return zf.read(target)


# ---------------------------------------------------------------------------
# 데이터 스토어 (스레드 안전)
# ---------------------------------------------------------------------------


class ToiletStore:
    """파싱된 레코드와 메타데이터를 보관하고 원자적으로 교체한다."""

    def __init__(self, cache_dir: Path = DEFAULT_CACHE_DIR):
        self.cache_dir = Path(cache_dir)
        self._lock = threading.RLock()
        self._records: list[dict] = []
        self._loaded_at: Optional[str] = None
        self._source: Optional[str] = None

    # --- 상태 ---
    @property
    def records(self) -> list[dict]:
        with self._lock:
            return self._records

    @property
    def count(self) -> int:
        with self._lock:
            return len(self._records)

    @property
    def loaded_at(self) -> Optional[str]:
        return self._loaded_at

    @property
    def source(self) -> Optional[str]:
        return self._source

    def _set(self, records: list[dict], source: str) -> None:
        with self._lock:
            self._records = records
            self._loaded_at = datetime.now(KST).isoformat(timespec="seconds")
            self._source = source

    # --- 적재 ---
    @property
    def _xlsx_path(self) -> Path:
        return self.cache_dir / INNER_XLSX_NAME

    def load_from_cache(self) -> bool:
        """캐시된 xlsx가 있으면 적재. 성공 시 True."""
        path = self._xlsx_path
        if not path.exists():
            return False
        records = parse_xlsx(path.read_bytes())
        self._set(records, f"cache:{path}")
        return True

    def refresh(self, timeout: int = 60) -> dict:
        """원격에서 다시 받아 파싱하고 캐시에 저장한다."""
        zip_bytes = download_zip(timeout=timeout)
        xlsx_bytes = extract_inner_xlsx(zip_bytes)
        records = parse_xlsx(xlsx_bytes)
        # 캐시 저장
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        (self.cache_dir / "contents.zip").write_bytes(zip_bytes)
        self._xlsx_path.write_bytes(xlsx_bytes)
        self._set(records, DOWNLOAD_URL)
        return self.info()

    def ensure_loaded(self, allow_download: bool = True, timeout: int = 60) -> dict:
        """데이터가 비어있으면 캐시->다운로드 순으로 적재."""
        if self.count > 0:
            return self.info()
        if self.load_from_cache():
            return self.info()
        if allow_download:
            return self.refresh(timeout=timeout)
        return self.info()

    # --- 메타 ---
    def info(self) -> dict:
        with self._lock:
            districts: dict[str, int] = {}
            for r in self._records:
                d = r.get("district")
                if d:
                    districts[d] = districts.get(d, 0) + 1
            return {
                "theme": THEME_NAME,
                "theme_id": THEME_ID,
                "count": len(self._records),
                "loaded_at": self._loaded_at,
                "source": self._source,
                "districts": dict(sorted(districts.items())),
                "source_url": DOWNLOAD_URL,
            }


# ---------------------------------------------------------------------------
# 질의 (스토어와 분리된 순수 함수)
# ---------------------------------------------------------------------------


def to_public(rec: dict, distance_m: Optional[float] = None, now: Optional[datetime] = None) -> dict:
    """LLM/클라이언트에 반환할 정제 딕셔너리."""
    oh = rec.get("open_hours_parsed") or {}
    out = {
        "content_id": rec.get("content_id"),
        "name": rec.get("name"),
        "subcategory": rec.get("subcategory"),
        "district": rec.get("district"),
        "road_address": rec.get("road_address"),
        "jibun_address": rec.get("jibun_address"),
        "latitude": rec.get("lat"),
        "longitude": rec.get("lon"),
        "phone": rec.get("phone"),
        "open_hours": rec.get("open_hours"),
        "open_now": is_open_now(oh, now),
        "is_24h": oh.get("is_24h", False),
        "closed_days": rec.get("closed_days", []),
        "toilet_types": rec.get("toilet_types", []),
        "disabled_toilet": rec.get("disabled_toilet", []),
        "amenities": rec.get("amenities", []),
        "safety": rec.get("safety", []),
        "building_use": rec.get("building_use"),
        "manager": rec.get("manager"),
        "open_hours_note": oh.get("note"),
    }
    if distance_m is not None:
        out["distance_m"] = round(distance_m, 1)
    return out


def _passes_filters(
    rec: dict,
    open_now: bool,
    require_disabled: bool,
    now: Optional[datetime],
) -> bool:
    if open_now:
        if is_open_now(rec.get("open_hours_parsed") or {}, now) is not True:
            return False
    if require_disabled:
        if not rec.get("disabled_toilet"):
            return False
    return True


def find_nearest(
    records: Iterable[dict],
    latitude: float,
    longitude: float,
    limit: int = 5,
    radius_m: Optional[float] = None,
    open_now: bool = False,
    require_disabled: bool = False,
    now: Optional[datetime] = None,
) -> list[dict]:
    scored: list[tuple[float, dict]] = []
    for rec in records:
        lat, lon = rec.get("lat"), rec.get("lon")
        if lat is None or lon is None:
            continue
        if not _passes_filters(rec, open_now, require_disabled, now):
            continue
        d = haversine_m(latitude, longitude, lat, lon)
        if radius_m is not None and d > radius_m:
            continue
        scored.append((d, rec))
    scored.sort(key=lambda x: x[0])
    return [to_public(rec, distance_m=d, now=now) for d, rec in scored[:limit]]


def search(
    records: Iterable[dict],
    query: Optional[str] = None,
    district: Optional[str] = None,
    open_now: bool = False,
    require_disabled: bool = False,
    limit: int = 20,
    now: Optional[datetime] = None,
) -> list[dict]:
    q = (query or "").strip().lower()
    dist = (district or "").strip()
    results: list[dict] = []
    for rec in records:
        if dist and dist not in (rec.get("district") or ""):
            continue
        if q:
            hay = " ".join(
                str(rec.get(f) or "")
                for f in ("name", "road_address", "jibun_address", "manager", "keyword")
            ).lower()
            if q not in hay:
                continue
        if not _passes_filters(rec, open_now, require_disabled, now):
            continue
        results.append(to_public(rec, now=now))
        if len(results) >= limit:
            break
    return results


def get_by_id(records: Iterable[dict], content_id: str) -> Optional[dict]:
    cid = (content_id or "").strip()
    for rec in records:
        if rec.get("content_id") == cid:
            pub = to_public(rec)
            pub["details"] = rec.get("details", {})  # 원문 상세 전체 포함
            return pub
    return None
